"""Scheduled report connections: pull from the systems you already have, drop informational
rows on the timeline (never tasks). A report source = a source row with Channel='report'
and ConfigJson {"type", "title", "every_minutes"/"daily_at", ...executor keys}.

REGISTRY: type -> executor(config) -> (headline, summary). Implemented: sqlite, rest, rss
(mssql with the [mssql] extra). Planned types fail loudly so a misconfig is visible on the
timeline instead of silently absent. Adding a type = one ~15-line function + a REGISTRY
entry - PRs welcome.
"""
import io, json, re, sqlite3, time
from datetime import datetime, timedelta
from loguru import logger

PLANNED = ['graphql', 'smb_file',
           # systems of record. Intacct is BUILT (see run_intacct); the rest are named because
           # the category is the question people arrive with - "does this reach our ERP / our
           # EMR" - and an empty Corporate systems group answers that worse than a list does.
           'netsuite', 'sap', 'workday', 'adp',            # quickbooks is BUILT (quickbooks.py)
           'epic', 'cerner', 'pointclickcare']   # smb_file is a NETWORK
# share and still planned; a path on this machine is local_file and works now

MAX_ROWS, BODY_CHARS, AI_CHARS = 200, 20000, 12000     # per report; override with cfg['max_rows']
SUMMARY_TOKENS = 1500     # a report summary is prose, not a triage verdict - give it room


def row_limit(cfg):
    """(limit, is it YOURS). The default is a safety net, not a number anybody chose - and
    the headline has to say which one it was, or "capped at 200" points the owner at a
    setting they never made and cannot find."""
    n = cfg.get('max_rows')
    return (max(1, int(n)), True) if n else (MAX_ROWS, False)


def rows_out(rows, limit, unit='rows', mine=True):
    """(headline, body) from executor rows, SAYING SO when the result was cut. A silent cap
    made the AI describe 20 rows of a TOP 500 query as 'all of them' - fetch one extra row
    and the headline can tell the truth instead."""
    more = len(rows) > limit
    rows = rows[:limit]
    why = (f'capped at {limit}' if mine else f'capped at the default {limit}') + ' — set "max rows" on this source to see more'
    head = f'{len(rows)} rows' + (f' ({why})' if more else '')
    return head.replace('rows', unit, 1), '\n'.join(json.dumps(r, default=str) for r in rows)[:BODY_CHARS]


def run_sqlite(cfg):
    """{"db": "path.db", "query": "SELECT ...", "max_rows": 200} - the local-first database report."""
    cx = sqlite3.connect(cfg['db']); cx.row_factory = sqlite3.Row
    lim, mine = row_limit(cfg)
    rows = [dict(r) for r in cx.execute(cfg['query']).fetchmany(lim + 1)]
    cx.close()
    return rows_out(rows, lim, mine=mine)


def run_mssql(cfg):
    """{"server", "database", "auth", "username", "password", "driver", "query", "max_rows"} -
    see mssql.py. Configure the connection entirely from the Connections tab."""
    from .mssql import run_report
    return run_report(cfg)


def run_intacct(cfg):
    """{"object": "GLENTRY", "fields": [...], "filters": [["BATCH_DATE", ">=", "08/01/2026"]],
    "max_rows": 200} - one readByQuery against Sage Intacct. The five credentials live on the
    Intacct card; a report carries only what it is asking for.

    Leave "fields" out and every field on the object comes back, which is the right default for
    a list somebody wants to eyeball and the wrong one for GL detail - so say which columns you
    want when the object is wide."""
    from .intacct import query
    obj = (cfg.get('object') or '').strip()
    if not obj: raise RuntimeError('no Intacct object set - e.g. GLENTRY, APBILL, VENDOR, LOCATION')
    lim, mine = row_limit(cfg)
    rows = query(cfg, obj, cfg.get('fields'), cfg.get('filters'), limit=lim + 1, order=cfg.get('order'))
    return rows_out(rows, lim, mine=mine)


def run_intacct_fields(cfg):
    """{"object": "APBILL"} - what the object actually HAS in this company, custom fields and
    all. It is a report in its own right (schedule it and a new custom field shows up on the
    timeline), and it is what the composer reads before writing an Intacct report."""
    from .intacct import fields_of
    obj = (cfg.get('object') or '').strip()
    if not obj: raise RuntimeError('no Intacct object set')
    lim, mine = row_limit(cfg)
    return rows_out(fields_of(cfg, obj), lim, unit='fields', mine=mine)


def run_metric(cfg):
    """{"name": "<metric>", "scope": "<what names one row>", "period": "2026-07"} - ONE certified number.

    This is the semantic layer's front door (semantic.py): the metric's definition was proved
    against numbers the owner already knew, so the answer is the company's number and not a
    plausible one. A metric that is not verified refuses rather than answering.
    """
    from . import semantic
    st = cfg.get('store')
    if st is None: raise RuntimeError('the metric tool reads the saved definitions - it needs the store')
    name = (cfg.get('name') or cfg.get('metric') or '').strip()
    if not name: raise RuntimeError('which metric? e.g. {"type": "metric", "name": "<metric>", "scope": "<what names one row>", "period": "2026-07"}')
    r = semantic.resolve(st, name, cfg.get('scope'), cfg.get('period'))
    head = f"{r['label']} · {r.get('scope') or 'all'} · {r.get('period') or 'all time'} = {r['value']:,.2f}"
    body = (f"{head}\n\n{r['definition']}\n\nVerified {r.get('verifiedAt') or ''} · {r['rows']} row(s) from "
            f"{r['object']} · filters {json.dumps(r['filters'])}")
    return head, body


def run_metric_check(cfg):
    """{"name": "<metric>"} or {} for all of them - re-prove the definitions against their known
    numbers. Scheduled, this is the tripwire: a chart-of-accounts change stops reconciling and
    the metric is demoted to broken on the timeline instead of quietly returning a wrong figure."""
    from . import semantic
    st = cfg.get('store')
    if st is None: raise RuntimeError('the metric check reads the saved definitions - it needs the store')
    name = (cfg.get('name') or '').strip()
    rows = [st.metric_by_name(name)] if name else st.list_metrics()
    if not rows or rows == [None]: raise RuntimeError(f'no metric called {name!r}' if name else 'no metrics defined yet')
    out, bad = [], 0
    for m in rows:
        r = semantic.check(st, m['MetricId'], cfg.get('actor') or 'schedule')
        bad += r['status'] != 'verified'
        out.append(f"{r['status'].upper():9} {r['name']} — {r['passed']}/{r['of']} known numbers reconcile"
                   + (f" · {r['note']}" if r['note'] else ''))
        out += [f"    {x['scope']} {x['period']}: expected {x['expected']:,.2f}, got "
                + (f"{x['got']:,.2f} (off {x['off']:,.2f})" if x.get('got') is not None else f"ERROR {x.get('error')}")
                for x in r['results'] if not x['pass']]
    head = (f"{len(rows) - bad}/{len(rows)} metric(s) still reconcile"
            + (f" · {bad} need attention" if bad else ''))
    return head, '\n'.join(out)


AGENT_SYSTEM = ('You are running a SCHEDULED REPORT for a busy operator. Do exactly what the instruction '
                'says - use your tools, read what you need to read - and then answer with the report itself: '
                'plain text or markdown, concrete (numbers, names, dates, deltas), no preamble and no '
                'questions back. If something the instruction asks about cannot be found, say so in the report.')


def _outline(body: str) -> str:
    """Section headings and table header rows of a report body - the shape without the content."""
    out, lines = [], (body or '').splitlines()
    for i, l in enumerate(lines):
        s = l.strip()
        if s.startswith('#'): out.append(s)
        elif s.startswith('|') and i + 1 < len(lines) and set(lines[i + 1].strip()) <= set('|-: ') and '-' in lines[i + 1]: out.append(s)
    return ' / '.join(out[:24])


def run_agent(cfg):
    """{"agent": "coder", "skill": "weekly-user-review", "prompt": "...", "cwd": "C:/repo", "model": "..."} -
    the AI itself as the source: a coding CLI agent (Connections -> AI CLI agents) runs your saved
    SKILL (a slash command - "/weekly-user-review") and/or a prompt, on the schedule, and what it
    answers is the report. "cwd" is optional: a project-level skill lives in its repo, a user-level
    one runs from anywhere. The AI summary pass is usually unnecessary - the agent already wrote prose.

    This is the "run my Claude skill every Monday" report: the agent researches, reads the systems
    it has tools for, and files what it found onto the Timeline like any other report."""
    from .llm import make_cli_llm
    store = cfg.get('store')
    if store is None: raise RuntimeError('the agent source needs the store (run it through the reports pipeline)')
    skill, prompt = str(cfg.get('skill') or '').strip().lstrip('/'), str(cfg.get('prompt') or '').strip()
    if not skill and not prompt: raise RuntimeError('give the agent a skill (/name) or a prompt - or both')
    name = str(cfg.get('agent') or 'coder').strip()
    llm = make_cli_llm(store, name, cfg.get('model') or None, cwd=cfg.get('cwd') or None)
    if llm is None: raise RuntimeError(f'no CLI agent named {name!r} - add one under Connections -> AI CLI agents')
    # Long workflows promoted from an assistant conversation live in Taskuary's neutral skill
    # store. Expand them into the prompt so one skill works through Claude, Codex, Gemini, or any
    # custom CLI. A skill that is not there remains the provider's normal `/skill-name` command.
    owned = None
    if skill:
        from . import config
        candidate = config.home() / 'skills' / skill / 'SKILL.md'
        try: owned = candidate.read_text(encoding='utf-8') if candidate.is_file() else None
        except OSError: owned = None
    ask = (f'TASKUARY SKILL /{skill}\n\n{owned}\n\nRUN INPUT\n{prompt}' if owned is not None
           else (f'/{skill}' + (' ' if prompt else '') if skill else '') + prompt)
    # Two runs twenty minutes apart came back as two different documents - 106 lines with a
    # fast-risers table, then 83 lines in another shape. A fresh agent has no memory of the last
    # run, so the last run's SHAPE (headings, table columns - never the content) rides along.
    prev = store.last_report(cfg.get('title') or '') if cfg.get('title') else None
    shape = _outline((prev or {}).get('BodyText'))
    if shape: ask += ('\n\nSTRUCTURE: keep the sections, their order and the table columns of the previous run of this '
                      f'report so runs stay comparable - change only the content. Previous outline: {shape}')
    out = str(llm(AGENT_SYSTEM, ask) or '').strip()
    if not out: raise RuntimeError(f'{name} answered nothing')
    what = f'/{skill}' if skill else 'a prompt'
    return f'{name} ran {what} - {len(out.splitlines())} lines', out[:BODY_CHARS]


def run_rest(cfg):
    """{"url", "headers", "path": "a.b"} - GET a JSON endpoint, dot-path into it.

    Through webguard, because the same executor is reachable from POST /api/tools/run: the URL
    can come from an AGENT, whose context is full of mail this codebase calls data and never
    instructions. See webguard for what a fetch to 169.254.169.254 would otherwise be."""
    from . import webguard
    r = webguard.get(cfg['url'], headers=cfg.get('headers') or {})
    r.raise_for_status()
    data = r.json()
    for k in (cfg.get('path') or '').split('.'):
        if k: data = data[int(k)] if isinstance(data, list) else data.get(k)
    return (f'{len(data)} items' if isinstance(data, list) else 'ok'), json.dumps(data, indent=1, default=str)[:BODY_CHARS]


def run_rss(cfg):
    """{"url"} - latest titles from an RSS/Atom feed. Guarded like run_rest: same reason."""
    from . import webguard
    xml = webguard.get(cfg['url']).text
    titles = re.findall(r'<title>(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?</title>', xml)[1:11]
    return f'{len(titles)} new items', '\n'.join(f'- {t}' for t in titles)[:4000]


def run_winrm(cfg):
    """{"host", "script"} - run PowerShell ON a remote Windows box (WinRM / PS remoting,
    your current Windows credentials) and report its output. A box you can RDP into is
    usually domain-joined and WinRM-reachable already; if not, run Enable-PSRemoting on
    it once (elevated)."""
    import subprocess
    host, script = cfg['host'], cfg['script']
    p = subprocess.run(['powershell', '-NoProfile', '-NonInteractive', '-Command',
                        f'Invoke-Command -ComputerName {host} -ScriptBlock {{ {script} }}'],
                       capture_output=True, text=True, encoding='utf-8', errors='replace', timeout=180)
    if p.returncode != 0: raise RuntimeError((p.stderr or p.stdout or 'remote run failed')[:500])
    out = (p.stdout or '').strip()
    return f'{len(out.splitlines())} lines from {host}', out[:BODY_CHARS]


def run_mcp(cfg):
    """{"cmd", "args", "tool", "tool_args"} - call any MCP server's tool. See mcp.py."""
    from .mcp import run_report
    return run_report(cfg)


def run_database(cfg):
    """{"query"} - ANY database by connection string (postgres/mysql/snowflake/... URLs via
    SQLAlchemy, raw ODBC strings via pyodbc). The string lives on the 'Any database'
    connector card; see db.py."""
    from .db import run_report
    return run_report(cfg)


def run_aws(cfg):
    """{"service", "operation", "params", "path"} - any boto3 call with the AWS card's keys."""
    from .aws import run_aws as _run
    return _run(cfg)


def run_s3(cfg):
    """{"bucket", "key" | "prefix"} - read an S3 object, or list under a prefix. See aws.py."""
    from .aws import run_s3_object
    return run_s3_object(cfg)


def run_cwlogs(cfg):
    """{"log_group", "pattern", "hours"} - grep a CloudWatch log group. See aws.py."""
    from .aws import run_cloudwatch_logs
    return run_cloudwatch_logs(cfg)


def run_azure(cfg):
    """{"path", "api_version"} - GET any Azure Resource Manager object. See azure.py."""
    from .azure import run_azure as _run
    return _run(cfg)


def run_azblob(cfg):
    """{"account", "container", "blob" | "prefix"} - read or list Azure blob storage."""
    from .azure import run_azure_blob
    return run_azure_blob(cfg)


def run_azlogs(cfg):
    """{"workspace_id", "query", "hours"} - KQL against a Log Analytics workspace."""
    from .azure import run_azure_logs
    return run_azure_logs(cfg)


def run_entra_users(cfg):
    """{"filter", "select"} - Entra ID people, over Graph on the Azure card's app. See azure.py."""
    from .azure import run_entra_users as _run
    return _run(cfg)


def run_entra_groups(cfg):
    """{"group"} - a group's transitive members, or every group when blank."""
    from .azure import run_entra_groups as _run
    return _run(cfg)


def run_entra_signins(cfg):
    """{"hours", "failed_only"} - Entra sign-in activity (needs P1/P2 + AuditLog.Read.All)."""
    from .azure import run_entra_signins as _run
    return _run(cfg)


def run_entra_licenses(cfg):
    """Licence SKUs with seats consumed vs spare - the unused-seat report."""
    from .azure import run_entra_licenses as _run
    return _run(cfg)


def run_prometheus(cfg):
    """{"query" (PromQL)} - an instant query; each series is a row of its labels + value.
    The base URL (and an optional bearer token) live on the Prometheus card."""
    import requests
    base = (cfg.get('base_url') or '').strip().rstrip('/')
    if not base: raise RuntimeError('no Prometheus base URL set - Connections → Prometheus')
    hdr = {'Authorization': f"Bearer {cfg['token']}"} if cfg.get('token') else {}
    r = requests.get(f'{base}/api/v1/query', params={'query': cfg['query']}, headers=hdr, timeout=30)
    r.raise_for_status()
    j = r.json()
    if j.get('status') != 'success': raise RuntimeError(f"prometheus: {j.get('error') or j}")
    rows = [{**(s.get('metric') or {}), 'value': (s.get('value') or [None, None])[1]}
            for s in (j.get('data') or {}).get('result') or []]
    lim, mine = row_limit(cfg)
    return rows_out(rows, lim, unit='series', mine=mine)


def run_datadog(cfg):
    """{"name" (optional filter)} - your monitors and their states, the at-a-glance health
    board. Keys live on the Datadog card (api key write-only + application key)."""
    import requests
    site = (cfg.get('site') or 'datadoghq.com').strip()
    params = {'name': cfg['name']} if cfg.get('name') else {}
    r = requests.get(f'https://api.{site}/api/v1/monitor', params=params, timeout=30,
                     headers={'DD-API-KEY': cfg.get('api_key') or '', 'DD-APPLICATION-KEY': cfg.get('app_key') or ''})
    if r.status_code in (401, 403): raise RuntimeError(f'Datadog said {r.status_code} - check the API key + application key')
    r.raise_for_status()
    rows = [{'name': m.get('name'), 'state': m.get('overall_state'), 'type': m.get('type'),
             'muted': bool((m.get('options') or {}).get('silenced')), 'modified': m.get('modified')}
            for m in r.json()]
    rows.sort(key=lambda m: {'Alert': 0, 'Warn': 1, 'No Data': 2}.get(m['state'], 3))   # trouble first
    lim, mine = row_limit(cfg)
    return rows_out(rows, lim, unit='monitors', mine=mine)


def run_digest(cfg):
    """{"days": 1} - Taskuary's own activity as the data: open work, finished work, pending
    reviews, fresh verdicts, who wrote how often. The Morning digest ships as a report ON
    PURPOSE: the brief lands on the Timeline like any report, its prompt is edited on the
    Reports tab, deleting the source turns it off - and it demonstrates how reports work
    using data every install already has. `store` arrives via resolve_cfg, never persisted."""
    from .digest import gather
    days = int(cfg.get('days') or 1)
    head = 'yesterday and today so far, distilled' if days == 1 else f'the last {days} days, distilled'
    return head, gather(cfg['store'], days)


def run_assistant(cfg):
    """The 'Assistant' report - the post on the Timeline (assistant.py). Scheduled and worded on the
    Reports tab like the Morning digest, but it does not file prose: run_report_source hands the
    due run to assistant.run, which posts ideas with buttons and state. This executor is what
    PREVIEW shows - the facts a run would hand the model. `store` arrives via resolve_cfg."""
    from .assistant import facts
    return 'what the assistant would read right now', facts(cfg['store'], cfg.get('watch_source_ids'), cfg.get('watch_sources'))


def run_automate(cfg):
    """{"days": 30} - Taskuary's own traffic as the data: what repeats often enough to
    automate, and the concrete policy/report/switch that would kill it. Ships seeded as
    the weekly 'Automation ideas' report; see toil.py. `store` arrives via resolve_cfg."""
    from .toil import gather
    days = int(cfg.get('days') or 30)
    return f'the last {days} days of repeated toil', gather(cfg['store'], days)


def _planned(name):
    def _fail(cfg): raise NotImplementedError(f"connector type '{name}' is on the roadmap - not implemented yet")
    return _fail


def _newest(path: str, by: str = 'newest'):
    """The file a scheduled report should read. A glob is the point, not a convenience: an export
    that lands as sales-2026-08-25.csv has a different name every morning, so a report naming one
    file exactly is a report that works for a day. The NEWEST match is what "the latest export"
    means. Also accepts a folder, and a plain path unchanged."""
    from glob import glob
    from pathlib import Path
    p = Path(path).expanduser()
    if any(ch in str(p) for ch in '*?['):
        files = [h for h in (Path(x) for x in glob(str(p))) if h.is_file()]
        if not files: raise RuntimeError(f'nothing matches {path} - no file to read')
        # mtime is what "the export that just arrived" means, and it is the right default. But a
        # file NAMED by its date is the case where mtime lies: re-copy last month's archive and it
        # becomes the newest file on disk while sales-2026-08-01.csv is plainly not the latest
        # sales. pick='name' takes the highest-sorting name instead, which for an ISO date IS the
        # latest. The headline always says which file was read, so a wrong guess is visible.
        if str(by or 'newest').lower() == 'name': return sorted(files)[-1]
        return max(files, key=lambda h: h.stat().st_mtime)
    if not p.exists(): raise RuntimeError(f'{p} does not exist on this machine')
    return p


def _rows_from_text(text: str, delim: str = None) -> list:
    """A delimited file as dicts, sniffing the separator when it is not given: exports arrive as
    csv, tsv and semicolon-separated depending on who produced them and in what locale."""
    import csv
    sample = text[:4000]
    if not delim:
        try: delim = csv.Sniffer().sniff(sample, delimiters=',;\t|').delimiter
        except csv.Error: delim = ','
    return [dict(r) for r in csv.DictReader(io.StringIO(text), delimiter=delim)]


def run_local_file(cfg):
    """{"path": "C:/exports/*.csv", "tail": 50, "sheet": "Sheet1"} - a file, folder or glob on
    THIS machine. Taskuary already runs on the owner's own computer, so the spreadsheet somebody
    drops in a folder every morning is a report source like any other - and the alternative was
    a WinRM script or nothing.

    csv/tsv/json/jsonl come back as rows; xlsx too where openpyxl is installed. Anything else is
    read as text and the LAST `tail` lines are shown, because a log's news is at the bottom.
    A folder lists what is in it, newest first, which answers "did today's export arrive?"."""
    import json as _json
    p = _newest(cfg['path'], cfg.get('pick'))
    lim, mine = row_limit(cfg)
    if p.is_dir():
        rows = sorted(({'name': f.name, 'bytes': f.stat().st_size,
                        'modified': datetime.fromtimestamp(f.stat().st_mtime).strftime('%Y-%m-%d %H:%M')}
                       for f in p.iterdir() if f.is_file()),
                      key=lambda r: r['modified'], reverse=True)
        return rows_out(rows, lim, unit=f'files in {p.name}', mine=mine)
    suffix = p.suffix.lower()
    if suffix == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError('reading .xlsx needs openpyxl - run: pip install openpyxl '
                               '(csv, tsv, json and text need nothing)')
        wb = load_workbook(p, read_only=True, data_only=True)      # data_only: values, not formulae
        ws = wb[cfg['sheet']] if cfg.get('sheet') else wb.active
        it = ws.iter_rows(values_only=True)
        head = [str(h) if h is not None else f'col{i}' for i, h in enumerate(next(it, []) or [])]
        rows = [dict(zip(head, [('' if v is None else v) for v in r])) for r in it]
        wb.close()
        return rows_out(rows, lim, unit=f'rows from {p.name}', mine=mine)
    text = p.read_text(encoding='utf-8', errors='replace')
    if suffix in ('.csv', '.tsv'):
        return rows_out(_rows_from_text(text, '\t' if suffix == '.tsv' else cfg.get('delimiter')),
                        lim, unit=f'rows from {p.name}', mine=mine)
    if suffix == '.jsonl':
        rows = [_json.loads(l) for l in text.splitlines() if l.strip()]
        return rows_out(rows, lim, unit=f'records from {p.name}', mine=mine)
    if suffix == '.json':
        data = _json.loads(text or 'null')
        if cfg.get('path_expr'):
            for k in str(cfg['path_expr']).split('.'):
                if k: data = data[int(k)] if isinstance(data, list) else (data or {}).get(k)
        if isinstance(data, list): return rows_out(data, lim, unit=f'records from {p.name}', mine=mine)
        return f'{p.name}', _json.dumps(data, indent=1, default=str)[:BODY_CHARS]
    lines = text.splitlines()
    try: tail = max(1, int(cfg.get('tail') or 50))
    except (TypeError, ValueError): tail = 50
    shown = lines[-tail:]
    head = f'{p.name} - last {len(shown)} of {len(lines)} lines'
    return head, '\n'.join(shown)[:BODY_CHARS]


def _research(name):
    def run(cfg):
        from . import research
        return getattr(research, f'run_{name}')(cfg)
    run.__doc__ = f'research.run_{name} - see taskuary/research.py'
    return run


def _calendar(cfg):
    from .calendar import run_calendar
    return run_calendar(cfg)


def _lazy(module, fn):
    """An executor that lives in its own module, imported on first use - and the composer's
    catalog still sees the real docstring (compose._keys_doc reads it off the wrapper)."""
    import importlib
    def run(cfg): return getattr(importlib.import_module(f'.{module}', __package__), fn)(cfg)
    try: run.__doc__ = getattr(importlib.import_module(f'.{module}', __package__), fn).__doc__
    except Exception: run.__doc__ = f'{module}.{fn}'
    return run

REGISTRY = {'sqlite': run_sqlite, 'mssql': run_mssql, 'database': run_database,
            # the web as a source: plain REST, a key on a card, nothing new in the exe
            'exa': _research('exa'), 'tavily': _research('tavily'),
            'firecrawl': _research('firecrawl'), 'reader': _research('reader'),
            'local_file': run_local_file,
            'aws': run_aws, 's3_object': run_s3, 'cloudwatch_logs': run_cwlogs,
            'azure': run_azure, 'azure_blob': run_azblob, 'azure_logs': run_azlogs,
            'entra_users': run_entra_users, 'entra_groups': run_entra_groups,
            'entra_signins': run_entra_signins, 'entra_licenses': run_entra_licenses,
            'prometheus': run_prometheus, 'datadog': run_datadog,
            'winrm': run_winrm, 'mcp': run_mcp, 'rest': run_rest,
            'intacct': run_intacct, 'intacct_fields': run_intacct_fields,
            # QuickBooks Online: three reads, and the first two WRITES a Corporate system has here -
            # a bill and a paid expense, gated by the card's scope (proposals below write)
            'quickbooks': _lazy('quickbooks', 'run_quickbooks'), 'quickbooks_vendors': _lazy('quickbooks', 'run_quickbooks_vendors'),
            'quickbooks_accounts': _lazy('quickbooks', 'run_quickbooks_accounts'),
            'quickbooks_bill': _lazy('quickbooks', 'run_quickbooks_bill'), 'quickbooks_expense': _lazy('quickbooks', 'run_quickbooks_expense'),
            # the bank and card feed (teller.py): where a transaction comes from before it becomes a bill
            'teller_accounts': _lazy('teller', 'run_teller_accounts'), 'teller_transactions': _lazy('teller', 'run_teller_transactions'),
            'teller_balances': _lazy('teller', 'run_teller_balances'),
            # the semantic layer over the ERP: a number that was PROVED, and the check that keeps it proved
            'metric': run_metric, 'metric_check': run_metric_check,
            'rss': run_rss, 'digest': run_digest, 'automate': run_automate, 'assistant': run_assistant,
            'calendar': _calendar,       # the owner's busy times, off the Outlook (and Google) cards - read-only
            'agent': run_agent,          # the AI itself: a saved skill or a prompt, run by a CLI agent on the schedule
            # files & sheets people already keep: a Google Sheet, a SharePoint list, a file in a library
            'google_sheets': _lazy('sheets', 'run_google_sheets'),
            'sharepoint_list': _lazy('sharepoint', 'run_sharepoint_list'), 'sharepoint_file': _lazy('sharepoint', 'run_sharepoint_file'),
            # the knowledge base: documents indexed in this store (knowledge.py) - searched, and refreshed on a schedule
            'kb_search': _lazy('knowledge', 'run_kb_search'), 'kb_reindex': _lazy('knowledge', 'run_kb_reindex'),
            'handbook_search': _lazy('handbook', 'run_handbook_search'), 'handbook_write': _lazy('handbook', 'run_handbook_write'),
            'handbook_vote': _lazy('handbook', 'run_handbook_vote'),
            **{n: _planned(n) for n in PLANNED}}

# Which connector CARD owns each executor type: the s3/cloudwatch types run on the aws
# card's keys, the blob/logs types on the azure card's app - roles and creds resolve there.
CARD_OF = {'s3_object': 'aws', 'cloudwatch_logs': 'aws', 'azure_blob': 'azure', 'azure_logs': 'azure', 'calendar': 'outlook',
           'entra_users': 'azure', 'entra_groups': 'azure', 'entra_signins': 'azure', 'entra_licenses': 'azure',
           'intacct_fields': 'intacct', 'sharepoint_list': 'sharepoint', 'sharepoint_file': 'sharepoint',
           'quickbooks_vendors': 'quickbooks', 'quickbooks_accounts': 'quickbooks', 'quickbooks_bill': 'quickbooks', 'quickbooks_expense': 'quickbooks',
           'teller_accounts': 'teller', 'teller_transactions': 'teller', 'teller_balances': 'teller',
           'kb_search': 'knowledge', 'kb_reindex': 'knowledge',
           'handbook_search': 'handbook', 'handbook_write': 'handbook', 'handbook_vote': 'handbook'}

def card_of(t): return CARD_OF.get(t, t)


def _connector(store, typ, connector_id=None, with_secret=False):
    """Resolve an explicitly selected instance, or the active/default instance for legacy
    report configs. Refuse an id belonging to another connector type."""
    c = store.get_connector(int(connector_id), with_secret=with_secret) if connector_id else \
        store.get_connector_by_type(typ, with_secret=with_secret)
    return c if c and c.get('Type') == typ else None


def mssql_connection(store, connector_id=None) -> dict:
    """The SQL Server CONNECTION lives on the mssql connector card (set up once, tested
    there); report configs carry only query/ai_prompt/schedule and inherit it here.
    Per-report overrides still win if present."""
    c = _connector(store, 'mssql', connector_id, with_secret=True)
    if not c: return {}
    cfg = json.loads(c.get('ConfigJson') or '{}')
    if c.get('Secret'): cfg.setdefault('password', c['Secret'])
    return {k: v for k, v in cfg.items() if v}


def winrm_connection(store, connector_id=None) -> dict:
    """Same connection-card pattern as mssql: the host lives on the winrm connector."""
    c = _connector(store, 'winrm', connector_id)
    cfg = json.loads((c or {}).get('ConfigJson') or '{}')
    return {k: v for k, v in cfg.items() if v}


def _card(store, typ, secret_as, connector_id=None):
    c = _connector(store, typ, connector_id, with_secret=True)
    if not c: return {}
    cfg = json.loads(c.get('ConfigJson') or '{}')
    if c.get('Secret'): cfg.setdefault(secret_as, c['Secret'])
    return {k: v for k, v in cfg.items() if v}


def database_connection(store, connector_id=None) -> dict:
    """The connection string lives on the 'Any database' card; its write-only secret fills
    the string's {password} placeholder."""
    return _card(store, 'database', 'password', connector_id)


def aws_connection(store, connector_id=None) -> dict:
    return _card(store, 'aws', 'secret_access_key', connector_id)


def azure_connection(store, connector_id=None) -> dict:
    """The Azure card's own app, else the Outlook connector's saved Graph app - one app
    registration can hold Graph permissions AND Azure RBAC roles, so the borrow is real."""
    cfg = _card(store, 'azure', 'client_secret', connector_id)
    if not (cfg.get('client_id') and cfg.get('client_secret')):
        cfg = {**_card(store, 'outlook', 'client_secret'), **cfg}
    return cfg


def intacct_connection(store, connector_id=None) -> dict:
    """Five credentials, of which exactly one is a secret worth hiding: the API USER's
    password. The sender pair identifies the integration and the company id names the tenant -
    neither is a password to this company's books, and burying them write-only would only mean
    nobody can ever check the sender id for a typo."""
    return _card(store, 'intacct', 'user_password', connector_id)


def _teller_connection(store, connector_id=None) -> dict:
    from .teller import connection
    return connection(store, connector_id)


def _quickbooks_connection(store, connector_id=None) -> dict:
    from .quickbooks import connection
    return connection(store, connector_id)


def prometheus_connection(store, connector_id=None) -> dict:
    """base_url (+ optional bearer token as the write-only secret) lives on the card."""
    return _card(store, 'prometheus', 'token', connector_id)


def datadog_connection(store, connector_id=None) -> dict:
    """site + application key on the card; the API key is the write-only secret."""
    return _card(store, 'datadog', 'api_key', connector_id)


def _sharepoint_connection(store, connector_id=None) -> dict:
    from .sharepoint import sharepoint_connection
    return sharepoint_connection(store, connector_id)


def _sheets_connection(store, connector_id=None) -> dict:
    from .sheets import google_sheets_connection
    return google_sheets_connection(store, connector_id)


def _apikey_card(typ):
    """A card whose whole configuration is one key: the secret arrives as `api_key`."""
    return lambda store, connector_id=None: _card(store, typ, 'api_key', connector_id)


CONNECTION_OF = {'mssql': mssql_connection, 'winrm': winrm_connection, 'database': database_connection,
                 'exa': _apikey_card('exa'), 'tavily': _apikey_card('tavily'),
                 'firecrawl': _apikey_card('firecrawl'), 'reader': _apikey_card('reader'),
                 'aws': aws_connection, 's3_object': aws_connection, 'cloudwatch_logs': aws_connection,
                 'azure': azure_connection, 'azure_blob': azure_connection, 'azure_logs': azure_connection,
                 'entra_users': azure_connection, 'entra_groups': azure_connection,
                 'entra_signins': azure_connection, 'entra_licenses': azure_connection,
                 'prometheus': prometheus_connection, 'datadog': datadog_connection,
                 'intacct': intacct_connection, 'intacct_fields': intacct_connection,
                 **{t: _quickbooks_connection for t in ('quickbooks', 'quickbooks_vendors', 'quickbooks_accounts', 'quickbooks_bill', 'quickbooks_expense')},
                 **{t: _teller_connection for t in ('teller_accounts', 'teller_transactions', 'teller_balances')},
                 # both borrow: SharePoint the Outlook tenant app, Sheets the Gmail card's Google client
                 'sharepoint_list': _sharepoint_connection, 'sharepoint_file': _sharepoint_connection,
                 'google_sheets': _sheets_connection}


def resolve_cfg(store, cfg: dict) -> dict:
    if cfg.get('type') in ('digest', 'automate', 'assistant', 'agent', 'calendar', 'kb_search', 'kb_reindex',
                           'metric', 'metric_check'):
        return {**cfg, 'store': store}   # their data IS the store (the agent's: its profile; the calendar's: the cards; the knowledge base: its index)
    conn = CONNECTION_OF.get(cfg.get('type'))
    if conn:
        saved = conn(store, cfg.get('connector_id')) if cfg.get('connector_id') else conn(store)
        return {**saved, **{k: v for k, v in cfg.items() if v not in (None, '')}}
    return cfg


AI_SYSTEM = ('You summarize scheduled report data for a busy operator. Follow the operator '
             'instruction exactly. Be concise and concrete: numbers, names, deltas. Plain text only. '
             'The data may be a CAPPED slice of a larger result (the headline says so, and the rows '
             'may be cut mid-way) - never describe a capped or truncated slice as complete, and say '
             'plainly when something the instruction asks about is not present in the rows you got.')

# The rows come back as a spreadsheet and a chart, and the model that just read every row knows
# which column is the measure better than a heuristic hunting for "all numeric" does.
CHART_SYSTEM = ('\n\nThe rows are also turned into a bar chart for the reader. If ONE column is a '
                'measure worth plotting, end your answer with a single line:\n'
                'CHART: <value column> | <label column> | <short chart title>\n'
                'Use the exact column names from the data. Omit the line entirely when the rows are '
                'not worth plotting (no measure, one row, or every value the same) - a chart of '
                'nothing is worse than no chart.')


def run_sources(store, subs: list):
    """Several sources feeding ONE report: each runs on its own connection and query, the
    bodies are stacked under labeled headers, and the AI pass downstream sees all of them
    at once. The same connection can appear twice with different queries. One source
    failing is reported in place - it never takes the whole report down."""
    heads, bodies = [], []
    for i, sub in enumerate(subs, 1):
        t = sub.get('type', 'rest')
        label = (sub.get('label') or '').strip() or f'{t} #{i}'
        try:
            head, body = REGISTRY[t](resolve_cfg(store, dict(sub)))
        except Exception as e:
            head, body = 'FAILED', f'error: {str(e)[:400]}'
            logger.warning(f'report source "{label}" failed: {e}')
        heads.append(f'{label}: {head}')
        bodies.append(f'=== {label} ({head}) ===\n{body}')
    return ' · '.join(heads)[:400], '\n\n'.join(bodies)[:BODY_CHARS]


def report_llm(store, cfg: dict, default_llm):
    """The brain THIS report asked for (cfg['ai_brain'] in /api/brains values, optional
    cfg['ai_model'] override) - a heavier model for the weekly review, the cheap tier for
    pings. Falls back to the caller's default (the triage brain) when unset or broken."""
    if not (cfg.get('ai_brain') or cfg.get('ai_model')): return default_llm
    from .llm import build_llm
    try: return build_llm(store, cfg.get('ai_brain') or None, cfg.get('ai_model') or None) or default_llm
    except Exception as e:
        logger.warning(f'report brain unavailable, using the default: {e}')
        return default_llm


def report_system(store, cfg: dict, charts: bool = False) -> str:
    """The system prompt of a report's AI pass. Rows from a database get the report summarizer;
    the Morning digest is the ASSISTANT speaking (digest.system: COUNSEL.md's voice and its honesty
    rules) - the owner (2026-08-30): a brief of counts is useless, make it the assistant's summary."""
    if cfg.get('type') == 'digest' or 'digest' in {s.get('type') for s in cfg.get('sources') or []}:
        from .digest import system
        return system(store)
    return AI_SYSTEM + (CHART_SYSTEM if charts else '')


def render_report(store, cfg: dict, llm=None):
    """Run the executor(s), then (optionally) the AI pass: cfg['ai_prompt'] + a configured
    AI connector turn raw rows into the summary that lands on the timeline. The report may
    name its own brain and model (report_llm); `llm` is the default it falls back to."""
    llm = report_llm(store, cfg, llm)
    subs = [s for s in (cfg.get('sources') or []) if s.get('type')]
    if subs:
        head, summary = run_sources(store, subs)
    else:
        cfg = resolve_cfg(store, cfg)
        head, summary = REGISTRY[cfg.get('type', 'rest')](cfg)
    if cfg.get('ai_prompt') and llm:
        try:
            data = summary[:AI_CHARS]
            if len(summary) > AI_CHARS: data += '\n…(data truncated here - later rows were NOT shown to you)'
            charts = str(store.get_settings().get('report_images_enabled') or '1') == '1'
            ai = (llm(report_system(store, cfg, charts),
                      f"Instruction: {cfg['ai_prompt']}\n\nData ({head}):\n{data}",
                      max_tokens=SUMMARY_TOKENS) or '').strip()
            # an empty answer used to file as a bare '--- raw data ---' wall, which reads
            # like the prompt was never run. Say what happened instead.
            if not ai:
                ai = ('(the model returned an empty summary - it may have spent its budget thinking. '
                      'Try a shorter prompt, or a non-reasoning model for report summaries.)')
            return head, f"{ai}\n\n--- raw data ---\n{summary[:4000]}"
        except Exception as e:
            logger.warning(f'AI summary failed for report: {e}')
            return head, f'(AI summary failed: {str(e)[:200]})\n\n{summary}'
    if cfg.get('ai_prompt') and not llm:
        return head, f'(AI prompt set, but no active AI connector - raw data below)\n\n{summary}'
    return head, summary


def _cron_field(spec: str, lo: int, hi: int) -> set:
    """One cron field -> the set of matching values. Supports * , - and /step (numeric only)."""
    out = set()
    for part in str(spec).split(','):
        part, step = (part.split('/', 1) + ['1'])[:2]
        step = int(step)
        if part.strip() in ('*', ''): rng = range(lo, hi + 1)
        elif '-' in part: a, b = part.split('-', 1); rng = range(int(a), int(b) + 1)
        else: v = int(part); rng = range(v, v + 1)
        out.update(x for x in rng if lo <= x <= hi)
        if any(x < lo or x > hi for x in rng): raise ValueError(f'{spec}: out of range {lo}-{hi}')
    return out


def cron_prev(expr: str, now: datetime):
    """The most recent minute matching a 5-field cron (min hour dom month dow) at or before
    `now`, scanning back up to 35 days - None when malformed or nothing matches. Vixie rule:
    dom and dow both restricted means EITHER may match. dow: 0 and 7 are Sunday."""
    try:
        parts = str(expr).split()
        if len(parts) != 5: return None
        mins, hrs, doms, mons, dows = (
            _cron_field(p, lo, hi) for p, (lo, hi) in zip(parts, ((0, 59), (0, 23), (1, 31), (1, 12), (0, 7))))
    except ValueError:
        return None
    dows = {0 if x == 7 else x for x in dows}
    dom_star, dow_star = parts[2] == '*', parts[4] == '*'
    t = now.replace(second=0, microsecond=0)
    for _ in range(35 * 24 * 60):
        cd = (t.weekday() + 1) % 7                        # python Mon=0 -> cron Sun=0
        day_ok = (t.day in doms if dow_star else cd in dows if dom_star
                  else (t.day in doms or cd in dows))
        if t.minute in mins and t.hour in hrs and t.month in mons and day_ok: return t
        t -= timedelta(minutes=1)
    return None


def _ran_today(last_polled) -> bool:
    try: return str(last_polled)[:10] == datetime.now().strftime('%Y-%m-%d')
    except (TypeError, ValueError): return False


def _ran_this_week(last_polled) -> bool:
    """A WEEKLY brief should also greet you when you open the app - but the same brief seven
    launches running is the noise once_per_day was invented to stop, one rung up."""
    try: return (datetime.now() - datetime.fromisoformat(str(last_polled)[:19].replace(' ', 'T'))).days < 7
    except (TypeError, ValueError): return False


def is_due(cfg: dict, last_polled, startup: bool = False) -> bool:
    # on_startup is local-first scheduling: the app is a window you open, so "when I open
    # it" is a real schedule. Due exactly once per launch - never on the 10-minute auto-sync,
    # and a cron time it would have missed while closed is not its problem.
    # on_startup ALONE is "once per launch, never on the clock"; on_startup beside a schedule is
    # BOTH - the Morning digest runs when the app opens and again on its interval
    # ...but a BRIEF is once a day. on_startup on the Morning digest filed a fresh copy on every
    # launch - ten identical briefs in two days, which is the noise that made it unreadable (the
    # owner, 2026-08-30). once_per_day keeps the "you opened the app, here is today's" behaviour
    # and drops every repeat: a launch fires it only when today has not had one yet.
    if cfg.get('on_startup'):
        stale = not ((cfg.get('once_per_day') and _ran_today(last_polled))
                     or (cfg.get('once_per_week') and _ran_this_week(last_polled)))
        if startup and stale: return True
        if not any(cfg.get(k) for k in ('cron', 'every_minutes', 'daily_at')): return False
    now = datetime.now()
    if not last_polled: return True
    try: last = datetime.fromisoformat(str(last_polled)[:19].replace(' ', 'T'))
    except ValueError: return True
    if cfg.get('cron'):
        # due when a scheduled minute passed since the last run. A local app sleeps: a cron
        # slot missed while closed fires on the next poll after reopening, once, not N times.
        prev = cron_prev(cfg['cron'], now)
        if prev is not None: return prev > last
        # malformed expression: fall through to the daily default, never a dead report
    if cfg.get('every_minutes'):
        try: return (now - last).total_seconds() >= float(cfg['every_minutes']) * 60
        except (TypeError, ValueError): pass               # 'every 30' typed as words: daily default
    if cfg.get('daily_at'):
        # tolerant of what people type: '8' and '8:30' both parse; garbage falls back to the
        # daily default instead of an unpack error killing the WHOLE poll thread (it did)
        try:
            hh, mm = (str(cfg['daily_at']).strip() + ':0').split(':')[:2]
            due = now.replace(hour=int(hh), minute=int(mm or 0), second=0, microsecond=0)
            return now >= due and last < due
        except (TypeError, ValueError):
            pass
    return (now - last).total_seconds() >= 24 * 3600


def findings_target(store, msg: dict) -> dict:
    """Where a report-born task's FINDINGS go when the agent is done, or {} for the default.

    The default is the Timeline and nowhere else, deliberately: a report is not a person, it is
    not waiting to hear anything, and mailing our own internal wrap-up back to whatever produced
    the numbers is the exact failure TQ-0252 was (an internal note posted to a robot's inbox).
    So when a report's task closes, the report lands on the task and stops there - unless the
    owner has said otherwise on that report's card, which is what deliver_findings is: a channel
    and an address they chose, for the one case where somebody genuinely does want telling."""
    if (msg or {}).get('Channel') != 'report': return {}
    cid = str(msg.get('ConversationId') or '')
    sid = cid.split(':', 1)[1] if cid.startswith('report:') else ''
    src = next((s for s in store.list_sources(active_only=False) if str(s.get('SourceId')) == sid), None)
    if not src: return {}
    try: cfg = json.loads(src.get('ConfigJson') or '{}')
    except ValueError: return {}
    d = cfg.get('deliver_findings') or {}
    if not (isinstance(d, dict) and d.get('to')): return {}
    from .outbound import can_reply
    ch = d.get('channel') or 'email'
    if not can_reply(store, ch): return {}
    return {'channel': ch, 'to': d['to'], 'subject': f"{cfg.get('title') or src.get('Address')} - what we found"}


LAST_RUN = 'report_last_run:'      # setting per source: what its last run did (the Reports tab shows it)


def last_runs(store) -> dict:
    """{source id: record} for every report that has run - when, how long, what it read, what came out."""
    out = {}
    for k, v in store.get_settings().items():
        if not k.startswith(LAST_RUN): continue
        try: out[int(k[len(LAST_RUN):])] = json.loads(v)
        except (ValueError, TypeError): continue
    return out


def run_report_source(store, src: dict, llm=None) -> dict:
    """Execute one due report and file it on the timeline - and leave a record of the run on the
    source (LAST_RUN): when, how long, what it read, what it reviewed, what it posted or why it
    stayed quiet. A quiet assistant check posts NOTHING, so without this there was no way to see
    what it had looked at (the owner, 2026-08-30: 'want to see the last run... what data is processed')."""
    t0, cfg = time.time(), json.loads(src.get('ConfigJson') or '{}')
    rec = {'at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'type': cfg.get('type') or 'rest', 'title': cfg.get('title') or src['Address']}
    def keep():
        # the newest run rides on the row (LAST_RUN); every run joins the history (report_run) - the
        # owner (2026-08-30): "a history of runs... to see what it processed and why it created certain things"
        store.set_setting(f"{LAST_RUN}{src['SourceId']}", json.dumps(rec, default=str), 'report')
        try: store.add_report_run(src['SourceId'], rec)
        except Exception as e: logger.warning(f'report run history not kept for {src["Address"]}: {e}')
    try:
        out = _run_report_source(store, src, cfg, llm)
    except Exception as e:
        rec.update({'ms': int((time.time() - t0) * 1000), 'failed': True, 'error': str(e)[:600]})
        keep(); raise
    rec.update({'ms': int((time.time() - t0) * 1000), 'subject': out.get('subject'), 'message_id': out.get('message_id'),
                'failed': str(out.get('subject') or '').endswith('FAILED'), 'files': out.get('files'),
                'said': out.get('said'), 'reviewed': out.get('reviewed'), 'inputs': str(out.get('inputs') or '')[:30000],
                'lines': out.get('lines') or [], 'summary': str(out.get('summary') or '')[:2000]})
    keep()
    return out


def _run_report_source(store, src: dict, cfg: dict, llm=None) -> dict:
    """Execute one due report (executor + optional AI pass) and file it on the timeline.
    Errors file visibly too."""
    title = cfg.get('title') or src['Address']
    logger.debug(f'report run: {title} ({cfg.get("type", "rest")}, ai={bool(cfg.get("ai_prompt"))})')
    if cfg.get('type') == 'assistant':
        # not a report row: the assistant posts its own kind of row (ideas with buttons and state),
        # on this report's schedule and with this report's prompt as its instruction
        from . import assistant
        out = assistant.run(cfg['store'] if cfg.get('store') else store, report_llm(store, cfg, llm), force=True, instruction=cfg.get('ai_prompt'))
        return {'message_id': out.get('message_id'), 'subject': f"{title} - {out.get('said', 0)} line(s)", 'files': 0, **out}
    try:
        head, summary = render_report(store, cfg, llm)
        subject, body = f'{title} — {head}', summary
    except Exception as e:
        subject, body = f'{title} — FAILED', f'Report error: {str(e)[:500]}'
        logger.warning(f'report {src["Address"]} failed: {e}')
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # the CHART: line is an instruction to Taskuary about what to draw, not prose for the reader:
    # artifacts reads it off `body`, and what gets filed is the summary without it
    from .artifacts import strip_directive
    failed = subject.endswith('— FAILED')
    if cfg.get('triage') and not failed:
        # the report is a MESSAGE like any other: triage reads it under TRIAGE.md, and a task is what
        # TRIAGE.md says - so an agent's research report can hand its findings to the coding agent.
        # A failed run is never work; it files with its error like before.
        #
        # watch_for is what makes that judgement possible. Without it the classifier is reading a
        # table of numbers with no idea which numbers would be bad, so every run came out fyi and
        # "a report can start work" was a switch that did nothing. It is the owner's own sentence
        # about why this report exists and what would count as off (triage.classify_intent's
        # `watch`), and it is the ONLY thing the report gets to say about its own verdict.
        from .ingest import ingest_message
        out = ingest_message(store, msg={'external_id': f'report:{src["SourceId"]}:{stamp}', 'channel': 'report',
                                         'subject': subject, 'body': strip_directive(body), 'from_name': title,
                                         'conversation_id': f'report:{src["SourceId"]}', 'sent_at': stamp,
                                         'source_link': cfg.get('link'), 'source_name': title,
                                         'watch_for': str(cfg.get('watch_for') or '').strip() or None}, llm=llm)
        mid = out.get('message_id')
        if not mid:
            mid = store.add_message({'TaskId': None, 'ExternalId': f'report:{src["SourceId"]}:{stamp}:feed', 'ConversationId': f'report:{src["SourceId"]}',
                                     'Channel': 'report', 'SourceName': title, 'Subject': subject, 'FromName': title, 'SentAt': stamp,
                                     'BodyText': strip_directive(body), 'SourceLink': cfg.get('link'), 'Status': 'feed'})
    else:
        mid = store.add_message({'TaskId': None, 'ExternalId': f'report:{src["SourceId"]}:{stamp}',
                                 'ConversationId': f'report:{src["SourceId"]}', 'Channel': 'report',
                                 'SourceName': title, 'Subject': subject, 'FromName': title,
                                 'SentAt': stamp, 'BodyText': strip_directive(body),
                                 'SourceLink': cfg.get('link'), 'Status': 'feed'})
        store.add_route(mid, None, 'feed', None,
                        'scheduled report - informational, never a task' + ('' if not cfg.get('triage') else ' (this run failed, so it was not triaged)'), [], 'report')
    # the rows are the report: hand back the spreadsheet to open and the chart to look at, not
    # just prose about them. Prose-only reports (an AI summary, a failure) produce neither.
    try:
        from .artifacts import attach_report_output
        made = attach_report_output(store, mid, title, body)
    except Exception as e:
        made = []
        logger.warning(f'report artifacts for {title} failed: {e}')
    store.audit('message', mid, 'report', 'report', 'agent', title)
    # ...and if the report is meant to LEAVE, this is where it turns around. Same run, same row
    # on the timeline - it just travels the other way, and says so.
    if cfg.get('deliver', {}).get('to'):
        try: deliver_report(store, src, cfg, subject, strip_directive(body))
        except Exception as e:
            logger.warning(f'outbound delivery for {title} failed: {e}')
            store.add_route(mid, None, 'feed', None, f'the report ran; sending it out failed: {str(e)[:200]}',
                            [], 'report')
    # ...and the alert, which is the opposite of delivery: it says nothing at all unless the
    # result trips the rule. A failure to SEND an alert is itself worth seeing on the timeline -
    # an alarm that quietly could not reach you is the worst of both worlds.
    if cfg.get('alert', {}).get('to'):
        try:
            why = alert_fires(cfg, subject.split('—', 1)[-1].strip(), strip_directive(body), failed)
            if why: send_alert(store, src, cfg, why, subject, strip_directive(body))
        except Exception as e:
            logger.warning(f'alert for {title} failed: {e}')
            store.add_route(mid, None, 'feed', None, f'the report ran; its alert could not be sent: {str(e)[:200]}',
                            [], 'report')
    # the digest report is ALSO what keeps DIGEST.md alive: one run, two homes - the Timeline
    # row you read in the morning, and the doc the Docs tab shows
    if 'digest' in {cfg.get('type'), *(s.get('type') for s in cfg.get('sources') or [])}:
        from .digest import HEADER
        store.save_doc('digest', f'{HEADER}_refreshed {stamp[:16]}_\n\n{strip_directive(body)}\n', 'digest')
    return {'message_id': mid, 'subject': subject, 'files': len(made), 'summary': strip_directive(body)}


# ── alerts: the report that only speaks up when something is wrong ──────────────────────
# A scheduled report tells you what IS. The thing you actually want to know is when what is
# stops matching what should be - "did the nightly job run in the last two hours, and if not,
# say so on my phone". Delivery sends every run and is therefore useless for that: a message
# that arrives whether or not anything is wrong is a message you stop reading.
#
# So an alert is a CONDITION on the result plus somewhere to send it. Silence is the normal
# outcome; an alert arriving means something to look at.
ALERT_WHEN = ('nothing_came_back', 'something_came_back', 'fewer_than', 'more_than',
              'contains', 'missing', 'failed')
_LEADING_COUNT = re.compile(r'\s*(\d[\d,]*)\b')


def result_count(head: str, body: str) -> int:
    """How many things the report found. Row executors say it in the headline ("0 rows",
    "12 rows (capped...)"); anything else is counted by non-blank lines, which is the honest
    reading of a prose result."""
    m = _LEADING_COUNT.match(str(head or ''))
    if m: return int(m.group(1).replace(',', ''))
    return len([ln for ln in str(body or '').splitlines() if ln.strip()])


def alert_fires(cfg: dict, head: str, body: str, failed: bool = False) -> str:
    """Should this run speak up, and in what words? '' means stay quiet.

    Returns the reason, so what lands on the phone says which rule tripped rather than just
    repeating the report.
    """
    a = cfg.get('alert') or {}
    when = str(a.get('when') or '').strip().lower()
    if not when or not str(a.get('to') or '').strip(): return ''
    if when not in ALERT_WHEN: raise ValueError(f'unknown alert condition {when!r} - one of {", ".join(ALERT_WHEN)}')
    # A failed run is its own alarm: whatever the rule was, the report could not answer it, and
    # "no rows" from a query that never ran is not the same fact as "no rows" from one that did.
    if failed: return 'the report failed to run' if when == 'failed' else f'the report failed to run, so "{when}" could not be judged'
    if when == 'failed': return ''
    n = result_count(head, body)
    text = str(a.get('text') or '')
    hay = f'{head}\n{body}'.lower()
    if when == 'nothing_came_back': return 'nothing came back' if n == 0 else ''
    if when == 'something_came_back': return f'{n} came back' if n > 0 else ''
    if when == 'fewer_than':
        want = float(a.get('count') or 0)
        return f'only {n} came back, expected at least {want:g}' if n < want else ''
    if when == 'more_than':
        want = float(a.get('count') or 0)
        return f'{n} came back, more than the {want:g} expected' if n > want else ''
    if when == 'contains': return f'the result mentions "{text}"' if text and text.lower() in hay else ''
    if when == 'missing': return f'the result never mentions "{text}"' if text and text.lower() not in hay else ''
    return ''


def send_alert(store, src: dict, cfg: dict, why: str, head: str, body: str) -> dict:
    """Put the alert on the owner's phone (or wherever they chose), and on the timeline.

    Unlike `deliver`, this does NOT default to the review gate. The review gate exists so
    Taskuary never writes to OTHER PEOPLE unasked; an alert is the owner telling themselves
    something is wrong, and one that waits in a queue for approval is not an alert. The channel
    still has to be switched on under Settings → Replies, so "Taskuary may write to WhatsApp"
    remains one decision the owner made once.
    """
    a = cfg.get('alert') or {}
    to = a['to'] if isinstance(a.get('to'), list) else [x.strip() for x in str(a.get('to') or '').split(',') if x.strip()]
    title = cfg.get('title') or src['Address']
    subj = (a.get('subject') or f'{title} — {why}').strip()
    note = str(a.get('note') or '').strip()
    text = '\n\n'.join(x for x in [f'{title}: {why}.', note, f'{head}', str(body or '')[:1500]] if x)
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ch = a.get('channel') or 'whatsapp'
    mid = store.add_message({
        'ExternalId': f'alert:{src["SourceId"]}:{stamp}', 'ConversationId': f'report:{src["SourceId"]}',
        'Channel': ch, 'SourceName': title, 'Subject': subj, 'FromName': 'Taskuary', 'SentAt': stamp,
        'BodyText': text, 'Direction': 'out', 'Status': 'sent'})
    from . import outbound
    sent = outbound.send_out(store, ch, to, subj, text)
    store.add_route(mid, None, 'send', None, f'alert - {why}; sent to {", ".join(to) or "nobody"} on {ch}', [], 'report')
    store.audit('message', mid, 'alert_sent', 'report', 'agent', {'to': to, 'channel': ch, 'why': why})
    return {'message_id': mid, 'why': why, 'sent': sent}


def deliver_report(store, src: dict, cfg: dict, subject: str, body: str) -> dict:
    """A report that goes OUT: to an address the owner chose, on a channel they picked, either
    after they have read it or straight away.

    `gate` is the whole point and it defaults to 'review'. Everything else in this app holds to
    "nothing sends without you", and a scheduled job that mails your customers on its own would
    be the one place that promise did not hold. Choosing 'auto' is the owner saying, once, that
    THIS report is safe to send unread - not a default they discover afterwards.

    Either way it lands on the timeline as an outbound row, so the funnel shows both directions.
    """
    d = cfg.get('deliver') or {}
    gate = str(d.get('gate') or 'review').lower()
    to = d.get('to') if isinstance(d.get('to'), list) else [x.strip() for x in str(d.get('to') or '').split(',') if x.strip()]
    subj = (d.get('subject') or subject or cfg.get('title') or 'Report').strip()
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    mid = store.add_message({
        'ExternalId': f'out:{src["SourceId"]}:{stamp}', 'ConversationId': f'report:{src["SourceId"]}',
        'Channel': d.get('channel') or 'email', 'SourceName': cfg.get('title') or src['Address'],
        'Subject': subj, 'FromName': 'Taskuary', 'SentAt': stamp, 'BodyText': body,
        'Direction': 'out', 'Status': 'draft' if gate == 'review' else 'sent'})
    who = ', '.join(to) or 'nobody yet'
    if gate == 'review':
        store.add_review({'MessageId': mid, 'Kind': 'outbound', 'Status': 'pending', 'DraftText': body,
                          'Reason': f'{cfg.get("title") or "report"} → {who}. Approve to send it.',
                          'Deliver': json.dumps({'channel': d.get('channel') or 'email', 'to': to, 'subject': subj})})
        store.add_route(mid, None, 'draft', None,
                        f'outbound report waiting for you - approve in Review and it goes to {who}', [], 'report')
        store.audit('message', mid, 'outbound_drafted', 'report', 'agent', {'to': to, 'channel': d.get('channel')})
        return {'gate': 'review', 'message_id': mid, 'to': to}
    from . import outbound
    sent = outbound.send_out(store, d.get('channel') or 'email', to, subj, body)
    store.add_route(mid, None, 'send', None,
                    f'sent automatically to {who} - this report is set to send without review', [], 'report')
    store.audit('message', mid, 'outbound_sent', 'report', 'agent', {'to': to, 'channel': sent.get('channel')})
    return {'gate': 'auto', 'message_id': mid, 'sent': sent}


def run_due_reports(store, startup: bool = False) -> int:
    from .llm import build_llm
    try: llm = build_llm(store)
    except Exception: llm = None
    n = 0
    for src in store.list_sources():
        if src['Channel'] != 'report': continue
        if is_due(json.loads(src.get('ConfigJson') or '{}'), src.get('LastPolledAt'), startup):
            run_report_source(store, src, llm)
            store.touch_source(src['SourceId'])
            n += 1
    return n
